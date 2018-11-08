# ---------------------------
# resourse consumtion calculation
# (C) 2018 by Herby
#
# you only need to set this variables from bash
#
#--> IYO
#export CLIENT_ID="...."
#export CLIENT_SECRET="....."
#
# --> date for the report period
#export START=$(date -d '10/01/2018 00:00:00' +"%s")
#export END=$(date -d '10/30/2018 23:59:59' +"%s")
#
#--> excel file names
#export XLS_FILE="consumption_2018_10.xlsx"
#export XLS_FILE_TPL="consumption_tpl.xlsx"
#
# then run the script
# 1. to get your Accounts
#python3 export_all_in_zip.py list
#
# 2. to generate the XLSX
#python3 export_all_in_zip.py <account_id>
#
# and set the URL

URL='https://ch-lug-dc01-001.gig.tech/restmachine/cloudapi/'
URL='https://at-vie-dc01-001.gig.tech/restmachine/cloudapi/'


import os
from pathlib import Path
import argparse
import capnp
import datetime
from os import environ
import sys
import shutil
from shutil import copyfile
import requests
import json
import openpyxl
import zipfile

temp_zipfile = 'constemp.zip'

def getKey(item):
    return item[0]

def getJWT():
    iyourl = 'https://itsyou.online/v1/oauth/access_token'

    params = {'grant_type': 'client_credentials',
    	      'client_id': os.environ["CLIENT_ID"],
    		  'response_type': 'id_token',
    		  'client_secret': os.environ["CLIENT_SECRET"]}
    resp = requests.post(iyourl, params=params)

    if resp.status_code != 200:
    	print ('error in JWT')
    	sys.exit(2)

    return resp.text

def getZPFile():
    #construct get for consumtion ZipFile
    ZURL= URL+'accounts/getConsumption?accountId='+account+'&start='+start_date+'&end='+end_date
    r = requests.get(ZURL, headers=headers, stream=True)
    with open(temp_zipfile, 'wb') as f:
        shutil.copyfileobj(r.raw, f)
    return

def getCSInfo(actcs):
    params = {'cloudspaceId': actcs}
    resp = requests.post(URL+"cloudspaces/get", headers=headers,params=params)
    css=resp.json()
    return (css['status'],css['name'],css['description'],css['externalnetworkip'])

def getAccountInfo(accid):
    params = {'accountId': accid}
    resp = requests.post(URL+"accounts/get", headers=headers,params=params)
    ac=resp.json()
    return (ac['status'],ac['name'],ac['resourceLimits'])

def getAccounts():
    resp = requests.post(URL+"accounts/list", headers=headers)
    ac=resp.json()
    print ('"Accessable Accounts for given rights')
    print ('On G8: ',URL)
    print ('----------------------------------------')
    for a in ac:
        print ("ID:",a['id']," Name:",a['name'])
    return


# get environ
xlsfile = environ["XLS_FILE"]
xlstpl = environ["XLS_FILE_TPL"]
start_date = environ["START"]
end_date = environ["END"]

# JWT
JWT = getJWT()
if JWT == None:
	print ('No JWT')
	sys.exit(1)

AUTH = "Bearer " + JWT
headers = {'Authorization': AUTH}

if len(sys.argv) == 1:
        print (' no command ')
        print ('----------------------------------------')
        print (' usage .....')
        print (sys.argv[0]," list")
        print (sys.argv[0],"<AccountID>")
        print ('----------------------------------------')
        sys.exit(1)

if sys.argv[1] == "list":
    getAccounts()
    sys.exit(0)

# check arguments!!! #######################
#if sys.argv[1] :
#    sys.exit(1)

account = sys.argv[1]

#init excel file from template
copyfile(xlstpl, xlsfile)

# init capnp, and fetch if not here
capnp.remove_import_hook()
resourcemonitoring = "resourcemonitoring.capnp"
rc_file = Path(resourcemonitoring)
if not rc_file.exists():
    r = requests.get("https://raw.githubusercontent.com/0-complexity/openvcloud/master/libs/CloudscalerLibcloud/CloudscalerLibcloud/schemas/resourcemonitoring.capnp?$RANDOM", stream=True)
    f = open(resourcemonitoring, 'w')
    f.write (r.text)
    f.close()
resources_capnp = capnp.load("resourcemonitoring.capnp")

#get the resourse consumption zip file
getZPFile()

wb = openpyxl.load_workbook(xlsfile)
ws = wb['a']

idx=8
ws['A'+str(idx)]= 'Account'
ws['B'+str(idx)]= 'date'
ws['C'+str(idx)]=  'Year'
ws['D'+str(idx)]= 'Month'
ws['E'+str(idx)]= 'Day'
ws['F'+str(idx)]= 'Hour'
ws['G'+str(idx)]= 'Cloud Space ID'
ws['H'+str(idx)]= 'Machine Count'
ws['I'+str(idx)]= 'Total Memory'
ws['J'+str(idx)]= 'Total VCPUs'
ws['K'+str(idx)]= 'Disk Size'
ws['L'+str(idx)]= 'Disk IOPS Read'
ws['M'+str(idx)]= 'Disk IOPS Write'
ws['N'+str(idx)]= 'NICs TX'
ws['O'+str(idx)]= 'NICs RX'
ws['P'+str(idx)]= 'Win'
ws['Q'+str(idx)]= 'Redhat'
ws['R'+str(idx)]= '----'

mi = []  # list of VMs
ci = []  # list of Cloudspaces

with zipfile.ZipFile(temp_zipfile) as zfile:
    for file in zfile.namelist():
        path_list = file.split(os.sep)
        account = path_list[0]
        year = path_list[1]
        month = path_list[2]
        day = path_list[3]
        hour = path_list[4]
        date = datetime.datetime(int(year),int(month),int(day),int(hour)).strftime('%s')
        account_obj = resources_capnp.Account.from_bytes(zfile.read(file))
        for i, cs in enumerate(account_obj.cloudspaces):
            cs_id = cs.cloudSpaceId
            machines = len(cs.machines)
            vcpus = 0
            mem = 0
            disksize = 0
            disk_iops_read = 0
            disk_iops_write = 0
            nics_tx = 0
            nics_rx = 0
            win = 0
            redhat = 0
            for machine in cs.machines:
                #if machine.status = "RUNNING"
                m_su = 0
                m_tu = 0
                m_nu = 0
                m_win = 0
                m_rh = 0
                if machine.id == 0:
                    continue
                vcpus += machine.vcpus
                mem += machine.mem
                #print ("cpuMinutes:",machine.cpuMinutes," status:",machine.status)
                #print (cs.cloudSpaceId,"  ",machine.id,"  ",machine.imageName)

                # count licenses
                # windows per vCPU
                if machine.imageName.find('Windows') >=0:
                    win += 1
                    m_win = machine.vcpus
                # redhat 1x lic if <=4 and if > 4 vCPU 2x lic
                if machine.imageName.find('Redhat') >=0:
                    redhat += 1
                    m_rh = 1
                    if machine.vcpus > 4:
                        m_rh = 1

                for disk in machine.disks:
                    disk_iops_read += disk.iopsRead
                    disk_iops_write += disk.iopsWrite
                    disksize += disk.size
                    m_su += disk.size
                    m_tu += (disk.iopsRead+disk.iopsWrite)
                for nic in machine.networks:
                    nics_tx += nic.tx
                    nics_rx += nic.rx
                    m_nu += (nic.tx+nic.rx)
                # list all used cloudspaces
                cc = cs.cloudSpaceId
                if not (cc in ci):
                    ci.append(cc)

                # list all machines
                mmm = (machine.id)
                m = [row[1] for row in mi]
                try:
                    midx = m.index(mmm)
                except Exception as e:
                    midx = -1
                if midx >= 0:
                    mi[midx][3] += machine.vcpus
                    mi[midx][4] += machine.mem
                    mi[midx][5] += m_su
                    mi[midx][6] += m_tu
                    mi[midx][7] += m_nu
                    mi[midx][8] += m_win
                    mi[midx][9] += m_rh
                else:
                    #print ("neu:",)
                    mmm = [cs.cloudSpaceId,machine.id,machine.imageName,machine.vcpus,machine.mem,m_su,m_tu,m_nu,m_win,m_rh]
                    mi.append(mmm)

            idx +=1

            ws['A'+str(idx)]=account
            ws['B'+str(idx)]=int(date)
            ws['C'+str(idx)]=year
            ws['D'+str(idx)]=month
            ws['E'+str(idx)]=day
            ws['F'+str(idx)]=hour
            ws['G'+str(idx)]=cs_id
            ws['H'+str(idx)]=machines
            ws['I'+str(idx)]=mem
            ws['J'+str(idx)]=vcpus
            ws['K'+str(idx)]= disksize
            ws['L'+str(idx)]= disk_iops_read
            ws['M'+str(idx)]= disk_iops_write
            ws['N'+str(idx)]= nics_tx
            ws['O'+str(idx)]= nics_rx
            ws['P'+str(idx)]= win
            ws['Q'+str(idx)]= 1
            ws['R'+str(idx)]= 1
            ws['S'+str(idx)]="=param!C3*I"+str(idx)
            ws['T'+str(idx)]="=param!C2*J"+str(idx)
            ws['U'+str(idx)]="=param!C4*K"+str(idx)
            ws['V'+str(idx)]="=param!C5*(L"+str(idx)+"+M"+str(idx)+")"
            ws['W'+str(idx)]="=param!C6*(N"+str(idx)+"+O"+str(idx)+")"

ws = wb['Summary']
ai = getAccountInfo(account)
ws['B5'] = ai[1]
i = 13
ci_sort = sorted(ci)
for cc in enumerate(ci_sort):
    actRow = str(i)
    ws['B'+actRow] = cc[1]
    c = getCSInfo(cc[1])

    ws['L'+actRow] = c[0]
    ws['M'+actRow] = c[1]
    ws['N'+actRow] = c[2]
    ws['O'+actRow] = c[3]
    i += 1

i=7
ws = wb['VMs']
mi_sort = sorted(mi,key=getKey)
cs = 0
for x in mi_sort:
    if cs != x[0]:
        i += 1
        cs = x[0]
    actRow = str(i)
    ws['A'+actRow] = x[0]
    ws['B'+actRow] = x[1]
    ws['C'+actRow] = x[2]
    ws['D'+actRow] = x[3]
    ws['E'+actRow] = x[4]
    ws['F'+actRow] = x[5]
    ws['G'+actRow] = x[6]
    ws['H'+actRow] = x[7]
    ws['I'+actRow] = x[8]
    ws['J'+actRow] = x[9]
    ws['L'+actRow] = "=param!C2*D"+actRow
    ws['M'+actRow] = "=param!C3*E"+actRow
    ws['N'+actRow] = "=param!C4*F"+actRow
    ws['O'+actRow] = "=param!C5*G"+actRow
    ws['P'+actRow] = "=param!C6*H"+actRow
    ws['Q'+actRow] = "=param!C7*I"+actRow
    ws['R'+actRow] = "=param!C8*J"+actRow
#    if x[2].find('Windows') >= 0:
#        ws['S'+actRow] = "=param!B7*D"+actRow
#    if x[2].find('Redhat') >= 0:
#        if x[]
#        ws['T'+actRow] = "=param!B8"
    ws['U'+actRow] = "=SUM(L"+actRow+":N"+actRow+")"
    ws['V'+actRow] = "=U"+actRow+"+Q"+actRow+"+R"+actRow
#
    i +=1
wb.save(xlsfile)
